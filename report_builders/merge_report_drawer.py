from .abstract_drawer import AbstractDrawer
from db.repositories import (
    program_repository,
    report_repository,
    tractor_repository,
    department_repository
)
import openpyxl as oxl
from typing import List, Dict


class MergeReportBuilder(AbstractDrawer):

    def draw(self):
        departments = department_repository.get_all()
        work_book = oxl.Workbook()
        work_book.remove(work_book.active)
        work_book.create_sheet(
            title='Статистика',
        )

        for department in departments:
            ws = work_book.create_sheet(
                department.name
            )

            programs = department_repository.get_active_programs_for_department(department.id)

            self.create_header(ws,programs)

            self.create_table_header(
                ws,
                {
                    'Модель трактора': 16,
                    '№ трактора': 20,
                    'Опытный узел': 56,
                    'Дата и время обращения': 18,
                    'Продолжительность контроля, м/ч': 24,
                    'Наработка, м/ч': 12,
                    'Дефект выявлен на м/ч': 24,
                    'ПЭ: Комментарий': 104,
                    'Граничная дата гарантии': 24
                },
                len(programs) + 3
            )

            self.create_table(ws, programs, len(programs) + 4)
                


                

        work_book.save('отчет.xlsx')



    def create_table_header(self, sheet, columns: Dict[str, int], start_row):
        sheet.row_dimensions[1].height = None 

        for i, (header, width) in enumerate(columns.items()):
            cell = sheet.cell(row=start_row, column=i+1)
            cell.value = header
            cell.font = oxl.styles.Font(bold=True)
            sheet.column_dimensions[oxl.utils.get_column_letter(i+1)].width = width
            cell.alignment = oxl.styles.Alignment(wrap_text=True, horizontal='center')

    def create_header(self, sheet, programs):
        sheet.cell(row=1, column=1).value = 'Название узла'
        sheet.cell(row=1, column=1).font = oxl.styles.Font(bold=True)
        sheet.cell(row=1, column=1).alignment = oxl.styles.Alignment(wrap_text=True, horizontal='center')

        sheet.cell(row=1, column=6).value = 'Количество тракторов'
        sheet.cell(row=1, column=6).font = oxl.styles.Font(bold=True)
        sheet.cell(row=1, column=6).alignment = oxl.styles.Alignment(wrap_text=True, horizontal='center')

        sheet.cell(row=1, column=7).fill = oxl.styles.PatternFill(start_color="F1F1F1", end_color="F1F1F1", fill_type="solid")

        sheet.cell(row=1, column=8).value = '- Программы, нароботавшие нужное количество часов'
        sheet.cell(row=1, column=8).font = oxl.styles.Font(bold=True)
        sheet.cell(row=1, column=8).alignment = oxl.styles.Alignment(wrap_text=True, horizontal='center')

        sheet.merge_cells('A1:E1')

        for i, program in enumerate(programs):
            tractors_count = len(program.tractors)

            if tractors_count > 0:
                sheet.cell(row=i+2, column=1).value = program.name
                sheet.cell(row=i+2, column=6).value = tractors_count
            sheet.merge_cells(f'A{i+2}:E{i+2}')

        


    def create_table(self, sheet, programs, start_row):
        
        for each in programs:
            for i, report in enumerate(each.reports):
                sheet.cell(row=i+start_row, column=1).value = report.tractor.model_name
                sheet.cell(row=i+start_row, column=2).value = report.tractor.serial_number
                sheet.cell(row=i+start_row, column=3).value = report.program.name

                sheet.cell(row=i+start_row, column=4).value = report.report_time
                sheet.cell(row=i+start_row, column=5).value = report.program.observation_duration
                sheet.cell(row=i+start_row, column=6).value = report.operating_hours

                sheet.cell(row=i+start_row, column=7).value = report.defect_detected_at_hours
                sheet.cell(row=i+start_row, column=8).value = report.comment
                sheet.cell(row=i+start_row, column=9).value = report.tractor.warranty_expire_date




