from datetime import date, datetime
from dataclasses import dataclass, field
from typing import Dict, List, Set, Iterator
import pandas as pd
from abc import ABC, abstractmethod
import openpyxl as oxl
from openpyxl.styles import Font


# --- Объекты --- #
class Program:
    def __init__(
            self, 
            name : str, 
            observation_duration : int):
        
        self.name = name
        self.observation_duration = observation_duration

class Tractor:
    def __init__(
            self, 
            model_name : str,
            serial_number : str, 
            warranty_expire_date : date,):
        
        self.model_name = model_name
        self.serial_number = serial_number
        self.warranty_expire_date = warranty_expire_date

class Report:
    def __init__(
            self, 
            comment: str, 
            program: Program,  
            tractor : Tractor, 
            report_opperating_hours : float,  
            operating_hours : float, 
            report_time = datetime):
        
        self.comment = comment # текст комментария
        self.program = program # Отчет по программе
        self.tractor = tractor # Трактор по которому обращение
        
        self.report_opperating_hours = report_opperating_hours # Дефект выявлен на report_date моточасов  
        self.operating_hours = operating_hours # Текущая наработка моточасов
        
        self.report_time = report_time # Дата и время обращения
        

class Department:
    def __init__(
            self, 
            name : str):
        
        self.name = name # Название бюро

class ProgramDepartment:
    def __init__(self, program : Program, department: Department):
        self.program = program
        self. department = department

class TractorProgram:
    def __init__(
            self, 
            tractor : Tractor, 
            program : Program):
        
        self.tractor = tractor
        self.program = program

# --- Хранилище объектов --- #
@dataclass
class Storage:
    programs: Dict[str, 'Program'] = field(default_factory=dict)
    tractors: Dict[str, 'Tractor'] = field(default_factory=dict)
    departments: Dict[str, 'Department'] = field(default_factory=dict)

    program_departments: List['ProgramDepartment'] = field(default_factory=list)
    tractor_programs: List['TractorProgram'] = field(default_factory=list)

    def programs_in_department(self, department : Department) -> Iterator[Program]:
        for relation in self.program_departments:
            if relation.department == department:
                yield relation.program
    
    def tractors_in_program(self, program : Program) -> Iterator[Tractor]:
        for relation in self.tractor_programs:
            if relation.program == program:
                yield relation.tractor


    


# --- Классы создатели --- #
class ProgramCreator:
    def __init__(self, storage: Storage):
        self.storage = storage

    def create(self, name: str, duration: int) -> Program:
        if name in self.storage.programs:
            print(f"Программа {name} уже существует")
        program = Program(name, duration)
        self.storage.programs[name] = program
        return program
    

class TractorCreator:
    def __init__(self, storage: Storage):
        self.storage = storage

    def create(self, model: str, serial: str, warranty_date: date) -> Tractor:
        if serial in self.storage.tractors:
            print(f"Трактор с серийным номером {serial} уже существует")
        tractor = Tractor(model, serial, warranty_date)
        self.storage.tractors[serial] = tractor
        return tractor

    def add_program(self, tractor_serial: str, program_name: str) -> TractorProgram:
        tractor = self.storage.tractors.get(tractor_serial)
        program = self.storage.programs.get(program_name)
        if not tractor or not program:
            pass
            #print("Трактор или программа не найдены")

        for relation in self.storage.tractor_programs:
            if relation.tractor == tractor and relation.program == program:
                print(f"Связь {tractor} -> {program} уже существует")
                return relation
        
        relation = TractorProgram(tractor, program)
        self.storage.tractor_programs.append(relation)
        return relation

class DepartmentCreator:
    def __init__(self, storage: Storage):
        self.storage = storage

    def create(self, name: str) -> Department:
        if name in self.storage.departments:
            pass
            # print(f"Отдел {name} уже существует")
        department = Department(name)
        self.storage.departments[name] = department
        return department

    def add_program(self, dept_name: str, program_name: str) -> ProgramDepartment:
        department = self.storage.departments.get(dept_name)
        program = self.storage.programs.get(program_name)
        if not department or not program:
            print("Отдел или программа не найдены")
            return None
        # Проверяем, существует ли уже такая связь
        for relation in self.storage.program_departments:
            if relation.program == program and relation.department == department:
                print(f"Связь {program_name} -> {dept_name} уже существует")
                return relation
        
        relation = ProgramDepartment(program, department)
        print('создали связь')
        self.storage.program_departments.append(relation)
        print('успешно добавил связь')
        return relation

# --- Стратегия получения данных --- #
class IDataStrategy(ABC):
    def __init__(self, storage: Storage):
        self.storage = storage

    @abstractmethod
    def load_programs(self) -> List[Dict]:
        pass
    
    @abstractmethod
    def load_tractors(self) -> List[Dict]:
        pass
    
    @abstractmethod
    def load_departments(self) -> List[Dict]:
        pass

class ExcelChecker:
    def check_file(self, file_path: str, required_columns: List[str]) -> bool:
        print(f'файл передан {file_path}')
        try:
            # Читаем только заголовки для экономии памяти
            df = pd.read_excel(file_path, nrows=0)
            # Проверяем наличие всех требуемых колонок
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                print(f"Отсутствующие колонки: {missing_columns}")
                return False
                
            return True
            
        except FileNotFoundError:
            print(f"Файл не найден: {file_path}")
            return False
        except Exception as e:
            raise ValueError(f"Ошибка при проверке файла: {str(e)}")

class ExcelStrategy(IDataStrategy):
    def __init__(self, web_file : str, bitrix_path : str, storage : Storage):
        super().__init__(storage)
        self.web_file = web_file
        self.bitrix_path = bitrix_path
        checker = ExcelChecker()
        web_correct = checker.check_file(web_file, ['Опытный узел', 'Модель трактора', '№ трактора', 'Граничная дата гарантии'])
        bitrix_correct = checker.check_file(bitrix_path, ['Название', 'Примечание'])
        if not web_correct or not bitrix_correct:
            print(
                "Неверный формат файлов. Проверьте наличие необходимых колонок в файлах"
            )
    

    
    def load_programs(self):
        worker = ProgramCreator(self.storage)

        df = pd.read_excel(self.bitrix_path)
        
        for i, row in df.iterrows():
            name = row['Название']
            duration = row['Примечание']
            a = worker.create(name, duration)
    
    def load_tractors(self):
        worker = TractorCreator(self.storage)

        df = pd.read_excel(self.web_file).dropna(subset=['Модель трактора', '№ трактора', 'Граничная дата гарантии', 'Опытный узел'])
        
        for i, row in df.iterrows():
            model = row['Модель трактора']
            serial_number = row['№ трактора']
            warranty_date = row['Граничная дата гарантии']
            program = row['Опытный узел']

            worker.create(model, serial_number, warranty_date)
            worker.add_program(serial_number, program[4:])

    def load_departments(self):
        worker = DepartmentCreator(self.storage)

        df = pd.read_excel(self.bitrix_path)

        for i, row in df.iterrows():
            program_name = row['Название']
            names = row['Теги'].split(', ')
            for name in names:
                worker.create(name)  
                worker.add_program(name, program_name)
        

# Создание отчета
class ReportBuilder():
    def __init__(self, storage: Storage, path_done):
        self.storage = storage
        self.path_done = path_done

    def __create_header(self, headers, ws, column_widths):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)  # Делаем текст жирным
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width



    def build(self):
        work_book = oxl.Workbook()
        work_book.remove(work_book.active)

        for department in self.storage.departments:
            work_book.create_sheet(title=department)
            ws = work_book[department]
        
            self.__create_header(
                headers = [
                    "Модель трактора",
                    "№ трактора", 
                    "Опытный узел",
                    "Дата и время обращения",
                    "Продолжительность контроля, м/ч",
                    "Наработка, м/ч",
                    "Дефект выявлен на м/ч",
                    "ПЭ: Комментарий",
                    "Граничная дата гарантии"
                ],
                ws = ws,
                column_widths = {
                    'A': 16,  # Модель трактора
                    'B': 24,   # № трактора
                    'C': 56,   # Опытный узел
                    'D': 24,   # Дата и время обращения
                    'E': 24,   # Продолжительность контроля
                    'F': 16,   # Наработка
                    'G': 24,   # Дефект выявлен на м/ч
                    'H': 104,   # ПЭ: Комментарий
                    'I': 16    # Граничная дата гарантии
                }
            )

            for each in self.storage.program_departments:
                print(each.department.name)
            
                
                    
        work_book.save('отчет.xlsx')

        


class Manager:
    def __init__(
            self,
            storage : Storage,
            data_strategy : IDataStrategy,
            report_builder : ReportBuilder
    ):
        self.storage = storage
        self.data_strategy = data_strategy
        self.report_builder = report_builder
        
    def run(self):
        if self.data_strategy:
            print('начали заполнение данных')
            self.data_strategy.load_programs()
            print('загрузили программы')
            print(len(list(self.storage.programs)))

            self.data_strategy.load_tractors()
            print('загрузили тракторы')
            self.data_strategy.load_departments()
            print('загрузили отделы')
            self.report_builder.build()
            print('сделали отчет')


if __name__ == "__main__":
    # переменные
    web_path = r"C:\Users\Aleksandr\Downloads\Отчет(архивный) (5).xlsx"
    bitrix_path = r"C:\Users\Aleksandr\Downloads\битрикс.xlsx"
    result_path = r"C:\Users\Aleksandr\Downloads\Отчет.xlsx"


    # Инициализация компонентов
    storage = Storage()
    strategy = ExcelStrategy(
        web_file = web_path,
        bitrix_path = bitrix_path,
        storage = storage,
    )
    report_builder = ReportBuilder(storage, result_path)

    # Инициализация менеджера
    manager = Manager(storage, strategy, report_builder)

    # Запуск
    manager.run()